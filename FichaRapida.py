import openpyxl 
from PIL import Image, ImageDraw, ImageFont
import locale

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

# Abrindo a planilha
workbook_dados = openpyxl.load_workbook('C:/Users/Thiago Dias/Desktop/Projeto FichaRapida/dadosclientes.xlsx')
Sheet1_dados = workbook_dados['dados']

# Definindo a fonte (arial)
font_nome_X = ImageFont.truetype('./arial.ttf', 30)
font_texto = ImageFont.truetype('./arial.ttf',30)
font_valor = ImageFont.truetype('./arial.ttf', 20)
font_email = ImageFont.truetype('./arial.ttf', 27)

# Funão para desenhar o X 
def desenhar_X(desenhar, coord):
    desenhar.text(coord,'X',fill='black',font=font_nome_X)

# Extraindo os dados da planilha 
for linha in (Sheet1_dados.iter_rows(min_row=2, max_row=Sheet1_dados.max_row,min_col=1,max_col=Sheet1_dados.max_column)):
    tipoVenda = linha[0].value
    razaoSocial = linha[1].value
    cnpj = linha[2].value
    inscricaoEst = linha[3].value
    emailXml = linha[4].value
    enderecoCobranca = linha[5].value
    cttFinan = linha[6].value
    dddFinan = linha[7].value
    telFinan = linha[8].value
    emailFinan = linha[9].value
    cttCompr = linha[10].value
    dddCompr = linha[11].value
    telCompr = linha[12].value
    emailCompr = linha[13].value
    numFunci = linha[14].value
    tempoImovel = linha[15].value
    fatMensal = linha[16].value
    tipoImovel = linha[17].value
    valorImovel = linha[18].value 

    fatFormatado = locale.currency(fatMensal, grouping=True)
    valorImovelFormatado = locale.currency(valorImovel, grouping=True)

# Abrindo imagem
    image = Image.open("C:/Users/Thiago Dias/Desktop/Projeto FichaRapida/Ficha Cadastral Preenchimento Eletrônico - SETEMBRO 2021-1.png")
    desenhar = ImageDraw.Draw(image)

# Preenche tipo de venda
    coordAvista = (193, 149)
    coordAprazo = (535, 149)
    coordAlteracao = (879, 149)
    coord120e180 = (1112, 149)
    coordInativo = (1355, 149)

    if tipoVenda and tipoVenda.strip().upper() == 'A VISTA':
        desenhar_X(desenhar, coordAvista)
    elif tipoVenda and tipoVenda.strip().upper() == 'A PRAZO':
        desenhar_X(desenhar, coordAprazo)
    elif tipoVenda and tipoVenda.strip().upper() == 'ALTERAÇÃO':
        desenhar_X(desenhar, coordAlteracao)
    elif tipoVenda and tipoVenda.strip().upper() == 'ENTRE 120 E 180':
        desenhar_X(desenhar, coord120e180)
    elif tipoVenda and tipoVenda.strip().upper() == 'INATIVO>180':
        desenhar_X(desenhar, coordInativo)

# Coordenada para A vista (193, 149)
# Coordenada para A prazo (535, 149)
# Coordenada para Alteração (879, 149)
# Coordenada para Entre 120 e 180 (1112, 149)
# Coordenada para Inativo > 180 (1355, 149)

# Preenche Razão Social
    desenhar.text((329, 295), razaoSocial,fill='black', font=font_texto)

# Preenche Cnpj
    desenhar.text((329, 360),str(cnpj),fill='black',font=font_texto)

# Preenche Inscrição Estadual
    desenhar.text((738, 360), str(inscricaoEst), fill='black', font=font_texto)

# Preenche email para envio do XML
    desenhar.text((1128, 404), emailXml, fill='black', font=font_texto)

# Preenche endereço de cobrança diferente de faturamento
    coordSim = (1227, 454)
    coordNao = (1015, 454)

    if enderecoCobranca and enderecoCobranca.strip().upper() == 'SIM':
        desenhar_X(desenhar, coordSim)
    elif enderecoCobranca and enderecoCobranca.strip().upper( )== 'NÃO':
        desenhar_X(desenhar, coordNao)

# Coordenadas para "não" (1227, 454)
# Coordenadas para "sim" (1015, 454)

# Preenche contato do financeiro
    desenhar.text((312, 500), cttFinan,fill='black',font=font_texto)

# Preenche DDD do financeiro
    desenhar.text((755, 502), str(dddFinan),fill='black',font=font_texto)

# Preenche telefone do financeiro
    desenhar.text((822, 502), telFinan,fill='black',font=font_texto)

# Preenche email do financeiro
    desenhar.text((1185, 506), emailFinan,fill='black',font=font_email)

# Preenche nome do comprador
    desenhar.text((312, 541), cttCompr,fill='black',font=font_texto)

# Preenche DDD do comprador 
    desenhar.text((755, 545), str(dddCompr),fill='black',font=font_texto)

# Preenche telefone do comprador
    desenhar.text((822, 545), telCompr,fill='black',font=font_texto)

# Preenche email do comprador 
    desenhar.text((1185, 545), emailCompr,fill='black',font=font_email)

# Preenche número de funcionários
    desenhar.text((325, 657), str(numFunci),fill='black',font=font_texto)

# Preenche tempo no imóvel
    desenhar.text((869, 657), str(tempoImovel),fill='black',font=font_texto)

# Preenche faturamento mensal
    desenhar.text((1339, 662), str(fatFormatado),fill='black',font=font_valor)

# Preenche tipo de imóvel
    coordProprio = (317,723)
    coordAlugado = (464,723)

    if tipoImovel and tipoImovel.strip().upper() == 'PRÓPRIO':
        desenhar_X(desenhar, coordProprio)
    elif tipoImovel and tipoImovel.strip().upper() == 'ALUGADO':
        desenhar_X(desenhar, coordAlugado)
    
# Coordenada para Imóvel Próprio (317,723)
# Coordenada para Imóvel Alugado (464, 723)

# Preenche Valor do Imóvel
    desenhar.text((869, 720), str(valorImovelFormatado),fill='black',font=font_valor)

# Diretório da imagem
    diretorioImagem = 'C:/Users/Thiago Dias/Desktop/Projeto FichaRapida/Fichas Preenchidas'

# Salva a imagem
    image.save(f'{diretorioImagem}/{razaoSocial} ficha cadastral.png')
